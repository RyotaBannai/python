import pathlib
import os
from openpyxl import Workbook
# check formulra "HEX2DEC" in FORMULAE >> True
from openpyxl.utils import FORMULAE
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
import csv


def read_csv():
    csv_dir = pathlib.Path(os.path.dirname(__file__))/'csv'
    return csv_dir


def read_output_dir():
    op_dir = pathlib.Path(os.path.dirname(__file__))/'output'
    return op_dir


if __name__ == "__main__":
    test_csv_file_path = read_csv() / 'Sacramentorealestatetransactions.csv'
    print(test_csv_file_path)
    output_file_name = 'test2.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = 'test excel file name'

    with open(test_csv_file_path, mode='r') as csv_file:
        mycsv = csv.reader(csv_file)
        for index, line in enumerate(mycsv):
            print(line)
            # [, , , ,...]
            address = line[0]
            price = line[9]
            new_entry = [address, price]
            ws.append(new_entry)
            alart_font = ''

            if price != 'price' and int(price) < 10000:
                alart_font = Font(color="FFBB00")
            else:
                alart_font = Font(color=RED)
            ws["B{}".format(index + 1)].font = alart_font

    wb.save(filename=read_output_dir()/output_file_name)

# ルールもセットできる
# >>> red_fill = PatternFill(bgColor="FFC7CE")
# >>> dxf = DifferentialStyle(fill=red_fill)
# >>> r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
# >>> r.formula = ['$A2="Microsoft"']
# >>> ws.conditional_formatting.add("A1:C10", r)
